from nsetools import Nse
import openpyxl
from datetime import datetime

# ✅ List of tickers to track (without .NS)
tickers = ['reliance', 'tcs', 'infy', 'axisbank', 'itc']

# ✅ Output Excel file
file_name = "nse_stock_prices.xlsx"

# Initialize NSE API
nse = Nse()

# Try to load workbook or create new
try:
    wb = openpyxl.load_workbook(file_name)
except FileNotFoundError:
    wb = openpyxl.Workbook()

# Use or create sheet
if "LivePrices" in wb.sheetnames:
    ws = wb["LivePrices"]
else:
    ws = wb.active
    ws.title = "LivePrices"
    ws.append(["Timestamp"] + [ticker.upper() for ticker in tickers])

# Fetch data
timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
prices = []

for ticker in tickers:
    try:
        stock = nse.get_quote(ticker)
        price = stock["lastPrice"]
    except Exception as e:
        price = "Error"
    prices.append(price)

# Write row to Excel
ws.append([timestamp] + prices)
wb.save(file_name)

print(f"✅ Saved prices to '{file_name}' at {timestamp}")
