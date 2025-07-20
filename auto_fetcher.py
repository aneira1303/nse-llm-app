import schedule
import time
from nsetools import Nse
import openpyxl
from datetime import datetime

# List of stock tickers to track (NSE format, lowercase)
tickers = ['reliance', 'infy', 'tcs', 'hdfcbank', 'itc']

# Excel file name
file_name = "nse_live_stock_data.xlsx"

# Initialize NSE object
nse = Nse()

def fetch_and_write_prices():
    print("🔄 Fetching stock prices...")

    # Load or create Excel workbook
    try:
        wb = openpyxl.load_workbook(file_name)
    except FileNotFoundError:
        wb = openpyxl.Workbook()

    # Use existing or new sheet
    if "Prices" in wb.sheetnames:
        ws = wb["Prices"]
    else:
        ws = wb.active
        ws.title = "Prices"
        ws.append(["Timestamp"] + [ticker.upper() for ticker in tickers])  # header

    # Fetch prices
    prices = []
    for ticker in tickers:
        try:
            stock = nse.get_quote(ticker)
            price = stock['lastPrice']
        except Exception:
            price = "Error"
        prices.append(price)

    # Append row with timestamp and prices
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.append([timestamp] + prices)

    wb.save(file_name)
    print(f"✅ Saved prices to Excel at {timestamp}")

# Run immediately once
fetch_and_write_prices()

# Schedule to run every 5 minutes
schedule.every(5).minutes.do(fetch_and_write_prices)

print("⏳ Scheduler running every 5 minutes... Press Ctrl+C to stop.")
while True:
    schedule.run_pending()
    time.sleep(1)
